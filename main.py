import tkinter as tk
import time
from datetime import datetime, date
from tkinter import messagebox, ttk, filedialog
from openpyxl import load_workbook
from MyQR import myqr
import os
import csv
import re

import uuid

import cv2
from pyzbar import pyzbar
from pyzbar.pyzbar import decode

#1st Page
root = tk.Tk()
root.title("Attendance Management Systen for IT Department")
root.geometry("600x500+320+45")
root.resizable(0, 0)
root.pack_propagate(False)
root['background'] = "lavender"

year = tk.StringVar()
sem = tk.StringVar()
batches = tk.StringVar()
subject = tk.StringVar()

def generate_qr():
    # Function to browse and select a folder
    def browse_folder():
        folder_path = filedialog.askdirectory()
        if folder_path:
            folder_path_entry.delete(0, tk.END)  # Clear any existing text in the Entry widget
            folder_path_entry.insert(0, folder_path)  # Insert the selected folder path

    # Function to browse and select a file
    def browse_file():
        file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            file_path_entry.delete(0, tk.END)  # Clear any existing text in the Entry widget
            file_path_entry.insert(0, file_path)  # Insert the selected file path

    # Function to generate QR codes
    def generate_qr_codes():
        folder_path = folder_path_entry.get()  # Get folder path from the entry widget

        # Create the folder if it doesn't exist
        os.makedirs(folder_path, exist_ok=True)

        file_path = file_path_entry.get()  # Get file path from the entry widget

        # Load the workbook
        try:
            workbook = load_workbook(file_path)

            # Select the first sheet
            sheet = workbook.active

            # Read the data from the sheet
            lines = []
            for row in sheet.iter_rows(values_only=True):
                line = ' '.join(str(cell) for cell in row)
                lines.append(line)

            for line in lines:
                if line:
                    data = line.strip()
                    save_name = f"{data}.png"
                    save_path = os.path.join(folder_path, save_name)

                    version, level, qr_name = myqr.run(
                        data,
                        version=1,
                        colorized=True,
                        contrast=1.0,
                        brightness=1.0,
                        save_name=save_name,
                        save_dir=folder_path
                    )
                    result_label.config(text=f"QR code generated for {data} and saved at {save_path}")
        except Exception as e:
            result_label.config(text=f"Error: {str(e)}")

    # Create the main window
    generate = tk.Tk()
    generate.title("QR Code Generator")
    generate.geometry('500x200+350+200')
    generate.pack_propagate(False)

    # Create and pack a frame for the input elements
    input_frame = ttk.Frame(generate, padding=10)
    input_frame.pack()

    # Label and Entry for specifying the folder path
    folder_path_label = ttk.Label(input_frame, text="Destination Folder:")
    folder_path_label.grid(row=1, column=0, sticky="w")

    folder_path_entry = ttk.Entry(input_frame, width=40)
    folder_path_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    # Button to browse for the destination folder
    browse_folder_button = ttk.Button(input_frame, text="Browse", command=browse_folder)
    browse_folder_button.grid(row=1, column=2, padx=5, pady=5, sticky="w")

    # Label and Entry for specifying the Excel file path
    file_path_label = ttk.Label(input_frame, text="Excel File Path:")
    file_path_label.grid(row=2, column=0, sticky="w")

    file_path_entry = ttk.Entry(input_frame, width=30)  # Reduced the width
    file_path_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")

    # Button to browse for the Excel file
    browse_file_button = ttk.Button(input_frame, text="Browse", command=browse_file)
    browse_file_button.grid(row=2, column=2, padx=5, pady=5, sticky="w")

    # Button to generate QR codes
    generate_button = ttk.Button(input_frame, text="Generate QR Codes", command=generate_qr_codes)
    generate_button.grid(row=3, columnspan=3, pady=10)

    # Label to display the result
    result_label = ttk.Label(generate, text="")
    result_label.pack()

generate_button = tk.Button(root, text="Generate QR Code", bg="lavender", fg="black", command=generate_qr, font=('times', 15, 'bold'))
generate_button.place(x=200, y=150)

def scan_qr():
    scan = tk.Toplevel(root)
    scan.title('Attendance Management System for IT Department')
    scan.geometry('600x500+320+45')
    scan['background'] = 'lavender'

    # Nested dictionary mapping options to related options
    related_options = {
        "1st": ["1st", "2nd"],
        "2nd": ["3rd", "4th"],
        "3rd": ["5th", "6th"],
        "4th": ["7th", "8th"]
    }


    def on_select(event):
        selected_option = combo_box.get()

        # Get the related options for the selected option
        if selected_option in related_options:
            options = related_options[selected_option]
        else:
            options = []

        # Clear the previous related options and populate the related combo box
        combo_box1['values'] = options

    # Create the combo box for selecting options
    label = tk.Label(scan, text="ScanIn", bg="lavender", fg="black", font=("times new roman", 25, 'bold'))
    label.place(x=230, y=17)

    label1 = tk.Label(scan, text="Year", bg="lavender", fg="black", font=("times new roman", 15))
    label1.place(x=120, y=100)
    combo_box = ttk.Combobox(scan, textvariable=year, values=["1st", "2nd", "3rd", "4th"], state="readonly")
    combo_box.bind("<<ComboboxSelected>>", on_select)
    combo_box.place(x=270, y=100)

    related_options1 = {
        "1st": ["EM-I", "Physics", "Graphics", "Communication Skills", "Energy and Environmental Engineering",
                "Basics civil and mechanical engineering"],
        "2nd": ["EM-II", "Chemistry", "Mechanics", "Computer programming in C",
                "Basics electrical and electronics engineering"],
        "3rd": ["EM-III", "Interpersonal Communication Skills", "Computer Architecture and Organization",
                "Object Oriented Programming in C++", "Data Structure and Applications"],
        "4th": ["Organization Behaviour", "Probability and Statistics", "Discrete Mathematics",
                "Design Analysis and Algorithms", "Digital Logic and Microprocessor", "Web Technology"],
        "5th": ["Software Engineering", "Computer Networks and Internetworking Protocols", "IT service management",
                "Network Management", "Data Visualization", "Virtual Reality", "Graph Theory", "Programming in Java",
                "Human Computer Interaction"],
        "6th": ["Operating System", "Database Management System", "Software testing",
                "Data Warehousing and Data Mining", "Compiler Design", "Enterprise Resource Planning",
                "Software Project Management", "Introduction to Data Science"],
        "7th": ["Cloud Computing and Storage Management", "Artificial Intelligence", "Pattern Recognition",
                "Soft Computing", "Electronic Payment System", "Natural Language Processing", "Machine Learning",
                "Real Time Systems", "Information Security", "Management Information Systems", "Distributed Computing"],
        "8th": ["Internet of Things", "Mobile Computing"]
    }

    def on_select1(event):
        selected_option1 = combo_box1.get()

        # Get the related options for the selected option
        if selected_option1 in related_options1:
            options = related_options1[selected_option1]
        else:
            options = []

        # Clear the previous related options and populate the related combo box
        combo_box2['values'] = options

    # Create the combo box for displaying related options
    label2 = tk.Label(scan, text="Semester", bg="lavender", fg="black", font=("times new roman", 15))
    label2.place(x=120, y=150)
    combo_box1 = ttk.Combobox(scan, textvariable=sem, state="readonly")
    combo_box1.bind("<<ComboboxSelected>>", on_select1)
    combo_box1.place(x=270, y=155)

    label4 = tk.Label(scan, text="Batch", bg="lavender", fg="black", font=("times new roman", 15))
    label4.place(x=120, y=205)
    combo_box3 = ttk.Combobox(scan, textvariable=batches, state="readonly")
    combo_box3['values'] = ['A',
                            'B',
                            'C',
                            'D']
    combo_box3.place(x=270, y=208)
    combo_box3.current()

    label3 = tk.Label(scan, text="Subject", bg="lavender", fg="black", font=("times new roman", 15))
    label3.place(x=120, y=260)
    combo_box2 = ttk.Combobox(scan, textvariable=subject, state="readonly")
    combo_box2.place(x=270, y=261)


    def check():
        if year.get() and sem.get() and batches.get() and subject.get():
            selected_year = year.get()  # Get the selected year from the combo box
            selected_semester = sem.get()  # Get the selected semester from the combo box
            selected_batch = batches.get()  # Get the selected batch from the combo box
            selected_subject = subject.get()  # Get the selected subject from the combo box


            # Store the year data in a variable or use it as needed
            print("Selected Year:", selected_year)
            print("Selected Semester:", selected_semester)
            print("Selected Batch:", selected_batch)
            print("Selected Subject:", selected_subject)
            scan.destroy()
        else:
            messagebox.showwarning("Warning", "All Fields are Required !!")

        '''Capturing'''
        capture = cv2.VideoCapture(0)
        names = []
        today = date.today()
        d = today.strftime("%b-%d-%Y")
        sub = subject.get()
        batch = batches.get()


        csv_filename = d + "_" + batch + "_" + sub + '.csv'
        file_exists = False
        try:
            with open(csv_filename, 'r') as csvfile:
                file_exists = True
        except FileNotFoundError:
            file_exists = False
        if not file_exists:
            # Create the CSV file and write the header row
            with open(csv_filename, 'w+', newline='') as csvfile:
                writer = csv.writer(csvfile, delimiter='\t')
                writer.writerow(["PR No. and Name \t\t", "Year\t", "Semester\t", "Subject\t", "In Time\t", "IP Address\t"])

        fob = open(csv_filename, 'a')  # Open the CSV file in append mode


        def enterData(z):
            if z in names:
                pass
            else:
                it = datetime.now()
                names.append(z)
                z = ''.join(str(z))
                intime = it.strftime("%H:%M:%S")

                # IP Address

                fob.write(z + '\t' + year.get() + '\t' + sem.get() + '\t' + batches.get() + '\t' + subject.get() + '\t' + intime + '\n')
            return names

        print('Reading...')

        def checkData(data):
            if data in names:
                print('Already Present')
            else:
                print('\n' + str(len(names) + 1) + '\n' + 'present...')
                enterData(data)

        while True:
            _, frame = capture.read()
            decodedObjects = pyzbar.decode(frame)
            for obj in decodedObjects:
                checkData(obj.data)
                time.sleep(0.5)

            gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            decoded = pyzbar.decode(gray)

            for barcode in decoded:
                x, y, w, h = barcode.rect
                barcode_data = barcode.data.decode('utf-8')
                barcode_type = barcode.type

                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.putText(frame, f'{barcode_data} ({barcode_type})', (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9,
                            (0, 255, 0),
                            2)

                # Process attendance data here
                print("QR Code Data:", barcode_data)

            cv2.imshow("Frame", frame)

            if cv2.waitKey(1) & 0xFF == ord('q'):
                cv2.destroyAllWindows()
                break

        root.withdraw()

        fob.close()


        # Read the CSV file and sort the data
        attendance_data = []
        with open(d + "_" + batch + "_" + sub + '.csv', 'r') as csvfile:
            reader = csv.reader(csvfile, delimiter='\t')
            headers = next(reader)
            attendance_data = sorted(reader, key=lambda x: int(re.sub(r'\D', '', x[0])))

        # Write the sorted data back to the CSV file
        with open(d + "_" + batch + "_" + sub + '.csv', 'w', newline='') as csvfile:
            writer = csv.writer(csvfile, delimiter='\t')
            writer.writerow(headers)
            writer.writerows(attendance_data)


        with open(d + "_" + batch + "_" + sub + '.csv', 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["PR No. and Name \t", "Year\t", "Semester\t", "Batch\t", "Subject\t", "In Time\t", "IP Address\t"])
            writer.writerows(attendance_data)


    # 2nd page
    button = tk.Button(scan, text="Submit", bg="lavender", fg="black", font=("times new roman", 16), command=check)
    button.place(x=250, y=330)


scan_button = tk.Button(root, text="Scan QR Code", bg="lavender", fg="black", command=scan_qr, font=('times', 15, 'bold'))
scan_button.place(x=200, y=230)

root.mainloop()
